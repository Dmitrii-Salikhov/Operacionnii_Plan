import pandas as pd
import logging
from datetime import datetime, timedelta, time as dt_time
from collections import defaultdict, Counter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from patient_parser import patient_parser, LOW_CONFIDENCE_THRESHOLD
from config_surgeons import pick_ma_surgeon
import config_surgeons
from constants import WEEKDAYS_RU, WEEKDAYS_FULL
from room_rules import (
    classify_calendar_title,
    extract_service_notes,
    is_service_event,
    is_tonsillectomy,
)

logger = logging.getLogger("plan_generator")


def parse_time_str(value, default=None):
    """Парсит 'HH:MM:SS' или 'HH:MM'. При ошибке — default или ValueError."""
    text = str(value or "").strip()
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            continue
    if default is not None:
        return default
    raise ValueError(f"Некорректное время: {value!r}")


# re-export для тестов / обратной совместимости
__all__ = [
    "OperationPlanGenerator",
    "parse_time_str",
    "is_service_event",
    "admissions_excel_filename",
]


def admissions_excel_filename(week_start):
    """Имя отдельного файла списка поступлений за неделю."""
    week_end = week_start + timedelta(days=6)
    return (
        f"Список поступлений ЛОР с {week_start.strftime('%d.%m.%Y')} "
        f"по {week_end.strftime('%d.%m.%Y')}.xlsx"
    )


class OperationPlanGenerator:
    def __init__(self, events_data=None, filepath=None, log_callback=None):
        self.log = log_callback if log_callback else lambda msg, tag='info': None
        if events_data is not None:
            self.df = pd.DataFrame(events_data)
        elif filepath is not None:
            self.df = self.load_all_sheets(filepath)
        else:
            raise ValueError("Нет данных для обработки")
        self.week_start = None
        self.events_by_day = defaultdict(list)
        self.daily_blocks = {i: {"5": [], "7": [], "MA": []} for i in range(5)}
        self.surname_counts = Counter()
        self.generalochka_days = set()

    def load_all_sheets(self, filepath):
        xls = pd.ExcelFile(filepath)
        dfs = []
        for sheet in xls.sheet_names:
            df = pd.read_excel(filepath, sheet_name=sheet)
            required = ['Название события', 'Дата начала (МСК)', 'Время начала (МСК)']
            if all(col in df.columns for col in required):
                dfs.append(df)
        if not dfs:
            raise ValueError("Ни один лист не содержит необходимых столбцов.")
        full_df = pd.concat(dfs, ignore_index=True)
        full_df.columns = [col.strip() for col in full_df.columns]
        return full_df

    def parse_all_events(self):
        for idx, row in self.df.iterrows():
            title = str(row.get('Название события', '')).strip()
            desc = str(row.get('Описание', '')).strip()
            date_str = str(row.get('Дата начала (МСК)', '')).strip()
            time_str = str(row.get('Время начала (МСК)', '')).strip()
            if not date_str or not title:
                continue

            # --- улучшенный парсинг даты с поддержкой разных форматов ---
            dt = None
            # пробуем явные форматы: YYYY-MM-DD, DD.MM.YYYY, ISO с часовым поясом
            for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%Y-%m-%dT%H:%M:%S%z', '%Y-%m-%dT%H:%M:%S'):
                try:
                    ts = pd.to_datetime(date_str, format=fmt, errors='coerce')
                    if not pd.isna(ts):
                        dt = ts.to_pydatetime()
                        break
                except (ValueError, TypeError, OverflowError, AttributeError) as e:
                    logger.debug("Формат даты %s не подошёл для %r: %s", fmt, date_str, e)
                    continue
            if dt is None:
                # последняя попытка – автоопределение с dayfirst=True
                ts = pd.to_datetime(date_str, dayfirst=True, errors='coerce')
                if not pd.isna(ts):
                    dt = ts.to_pydatetime()
            if dt is None:
                self.log(f"Некорректная дата '{date_str}' – событие '{title}' пропущено", 'warning')
                continue
            # -----------------------------------------------------------

            if self.week_start is None:
                self.week_start = dt - timedelta(days=dt.weekday())
            day_idx = dt.weekday()

            event_kind = classify_calendar_title(title)
            if event_kind == "holiday":
                continue
            if event_kind == "narcosis_closed":
                self.events_by_day[day_idx].append(
                    {"type": "narcosis_closed", "time": time_str, "date": dt}
                )
                continue
            if event_kind == "generalochka":
                self.events_by_day[day_idx].append(
                    {"type": "generalochka", "time": time_str, "date": dt}
                )
                self.generalochka_days.add(day_idx)
                continue
            if event_kind == "service_note":
                _, note_text = extract_service_notes(title)
                self.events_by_day[day_idx].append(
                    {
                        "type": "service_note",
                        "text": note_text or title.strip(),
                        "time": time_str,
                        "date": dt,
                    }
                )
                self.log(f"Служебная пометка: {title}")
                continue
            if event_kind == "service":
                self.log(f"Служебное событие пропущено: {title}")
                continue

            cleaned_title, title_notes = extract_service_notes(title)
            cleaned_desc, desc_notes = extract_service_notes(desc)
            # Пометки в скобках календаря обычно относятся к «Примечаниям» на Поступлении
            # (кроме скобок, где внутри есть ключ диагноза — их парсер раскрывает в diagnosis_raw).
            # Важно: делаем до extract_bound_notes, чтобы парсер не “съел” ключ внутри скобок.
            bracket_title_notes = patient_parser.extract_bracket_notes(cleaned_title)
            bracket_desc_notes = patient_parser.extract_bracket_notes(cleaned_desc)

            cleaned_title, bound_title = patient_parser.extract_bound_notes(cleaned_title)
            cleaned_desc, bound_desc = patient_parser.extract_bound_notes(cleaned_desc)
            patient = patient_parser.parse_patient_from_event(
                cleaned_title, cleaned_desc, logger=self.log
            )
            if patient is None:
                continue
            patient["date"] = dt
            patient["time"] = time_str
            notes_parts = [
                n
                for n in (
                    title_notes,
                    bracket_title_notes,
                    desc_notes,
                    bracket_desc_notes,
                    bound_title,
                    bound_desc,
                )
                if n
            ]
            if notes_parts:
                # без дублей при совпадении title/desc
                merged: list[str] = []
                for part in notes_parts:
                    for piece in part.split("; "):
                        piece = piece.strip()
                        if piece and piece not in merged:
                            merged.append(piece)
                patient["notes"] = "; ".join(merged)
            else:
                patient.setdefault("notes", "")
            self.events_by_day[day_idx].append(patient)

    def distribute_patients(self):
        for day in range(5):
            events = self.events_by_day.get(day, [])
            is_generalochka = day in self.generalochka_days
            narcosis_closed_time = None
            for ev in events:
                if ev.get("type") == "narcosis_closed":
                    t = ev.get("time", "23:59")
                    try:
                        narcosis_closed_time = parse_time_str(t)
                    except ValueError as e:
                        narcosis_closed_time = dt_time(23, 59)
                        self.log(
                            f"Некорректное время «закрыто для наркоза» ({t}): {e}. Принято 23:59",
                            "warning",
                        )
                    break

            patients = [
                p
                for p in events
                if p.get("type") not in ("narcosis_closed", "generalochka", "service_note")
            ]
            for p in patients:
                resolved = patient_parser.resolve_diagnosis(p['diagnosis_raw'])
                diag, operation = resolved['diagnosis'], resolved['operation']
                confidence = resolved['confidence']
                age, unit = p['age'], p['age_unit']
                if age is None:
                    age, unit = patient_parser.resolve_age_defaults(p['diagnosis_raw'], None)
                    self.log(f"Возраст {p['name']} не указан, принят {age} {unit}.")
                p['age'] = age
                p['age_unit'] = unit
                p['diagnosis'] = diag
                if operation == "Септопластика" and p.get('has_osteotomy'):
                    operation = "Септопластика (остеотомия)"
                p['operation'] = operation
                p['confidence'] = confidence
                p['confidence_source'] = resolved['source']
                note = (resolved.get("note") or "").strip()
                if note:
                    self._append_patient_note(p, note)
                # Низкая уверенность, неизвестный ключ или короткое ФИО → уточнение в GUI
                if (
                    p.get('is_unknown_diag')
                    or confidence < LOW_CONFIDENCE_THRESHOLD
                    or p.get('needs_name_review')
                ):
                    if p.get('is_unknown_diag') or confidence < LOW_CONFIDENCE_THRESHOLD:
                        p['is_unknown_diag'] = True
                    if confidence < LOW_CONFIDENCE_THRESHOLD and resolved['source'] != 'unknown':
                        self.log(
                            f"{p['name']} – низкая уверенность диагноза "
                            f"({confidence:.0%}, {resolved['source']}): "
                            f"«{p['diagnosis_raw']}» → {operation}",
                            'warning',
                        )
                    if p.get('needs_name_review'):
                        self.log(
                            f"Короткое ФИО «{p['name']}» — на уточнение.",
                            'warning',
                        )

                if is_generalochka:
                    if is_tonsillectomy(p['diagnosis_raw']):
                        self.log(f"{p['name']} – генералочка, тонзилэктомия отменена.")
                        continue
                    op_room = "MA"
                else:
                    is_ma = p['is_ma']
                    if is_ma:
                        if is_tonsillectomy(p['diagnosis_raw']):
                            op_room = "7"
                        else:
                            op_room = "MA"
                    else:
                        op_room = "5"

                    if op_room == "5" and narcosis_closed_time and p.get('time'):
                        try:
                            pt = parse_time_str(p['time'])
                        except ValueError as e:
                            self.log(
                                f"{p['name']} – некорректное время операции ({p.get('time')}): {e}",
                                "warning",
                            )
                            self.daily_blocks[day][op_room].append(p)
                            continue
                        if pt >= narcosis_closed_time:
                            if is_tonsillectomy(p['diagnosis_raw']):
                                op_room = "7"
                            else:
                                op_room = "MA"
                            p['is_ma'] = True
                            self.log(f"{p['name']} – после «Закрыто для наркоза» переведён в М/А.")

                self.daily_blocks[day][op_room].append(p)

            self._attach_day_service_notes(day)

        surnames = []
        for d in range(5):
            for room in ["5", "7", "MA"]:
                for p in self.daily_blocks[d][room]:
                    surname, _ = patient_parser.get_surname_and_initials(p['name'])
                    surnames.append(surname)
        self.surname_counts = Counter(surnames)

    @staticmethod
    def _time_to_minutes(value):
        try:
            tm = parse_time_str(value, default=None)
        except (TypeError, ValueError):
            return None
        if tm is None:
            return None
        return tm.hour * 60 + tm.minute

    @staticmethod
    def _append_patient_note(patient, text: str) -> None:
        note = (text or "").strip()
        if not note:
            return
        existing = (patient.get("notes") or "").strip()
        if not existing:
            patient["notes"] = note
            return
        parts = [p.strip() for p in existing.split(";") if p.strip()]
        if note not in parts:
            parts.append(note)
            patient["notes"] = "; ".join(parts)

    def _attach_day_service_notes(self, day: int) -> None:
        """Отдельные пометки календаря → ближайший пациент того же дня по времени."""
        notes = [
            e
            for e in self.events_by_day.get(day, [])
            if e.get("type") == "service_note"
        ]
        if not notes:
            return
        patients = []
        for room in ("5", "7", "MA"):
            patients.extend(self.daily_blocks[day][room])
        if not patients:
            for note in notes:
                self.log(
                    f"Пометка без пациента: {note.get('text')}",
                    "warning",
                )
            return
        for note in notes:
            note_mins = self._time_to_minutes(note.get("time"))
            best = None
            best_dist = None
            for p in patients:
                p_mins = self._time_to_minutes(p.get("time"))
                if note_mins is None or p_mins is None:
                    dist = 0 if best is None else 10**9
                else:
                    dist = abs(p_mins - note_mins)
                if best is None or best_dist is None or dist < best_dist:
                    best = p
                    best_dist = dist
            if best is None:
                continue
            text = note.get("text") or ""
            self._append_patient_note(best, text)
            self.log(f"Пометка «{text}» → {best.get('name')}")

    def assign_surgeons(self):
        # Читаем актуальные значения из модуля (после surgeons.save в UI).
        surgeon_5_map = config_surgeons.SURGEON_5
        surgeon_7 = config_surgeons.SURGEON_7
        surgeon_ma_map = config_surgeons.SURGEON_MA
        forbidden = config_surgeons.FORBIDDEN_MA
        surgeon_schedule = defaultdict(lambda: defaultdict(list))
        for day in range(5):
            blocks = self.daily_blocks[day]
            surgeon_5 = surgeon_5_map[day]
            for p in blocks["5"]:
                p['surgeon'] = surgeon_5
                surgeon_schedule[day][surgeon_5].append('5')
            for p in blocks["7"]:
                p['surgeon'] = surgeon_7
                surgeon_schedule[day][surgeon_7].append('7')
            ma_surgeon = pick_ma_surgeon(
                day, surgeon_5_map, surgeon_7, surgeon_ma_map, forbidden
            )
            for p in blocks["MA"]:
                p['surgeon'] = ma_surgeon
                surgeon_schedule[day][ma_surgeon].append('MA')

        for day, data in surgeon_schedule.items():
            for surg, rooms in data.items():
                if len(rooms) > 1:
                    self.log(f"⚠️ Конфликт в {WEEKDAYS_FULL[day]}: хирург {surg} назначен в {', '.join(rooms)}", 'warning')

    def sort_patients_in_rooms(self):
        for day in range(5):
            for room in ["5", "7", "MA"]:
                patients = self.daily_blocks[day][room]
                children = [p for p in patients if p['age'] is not None and p['age'] < 18]
                adults = [p for p in patients if p['age'] is not None and p['age'] >= 18]
                unknown = [p for p in patients if p['age'] is None]
                children.sort(key=lambda x: x['age'])
                adults.sort(key=lambda x: x['age'])
                self.daily_blocks[day][room] = children + adults + unknown

    def generate_excel(self, output_path):
        wb = Workbook()
        wb.remove(wb.active)

        header_font = Font(bold=True, size=11, name='Calibri')
        data_font = Font(size=10, name='Calibri')
        room_font = Font(size=30, name='Calibri')
        header_bottom_border = Border(bottom=Side(style='thin'))
        wrap_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        col_widths = {
            'A': 12, 'B': 8, 'C': 26, 'D': 36,
            'E': 30, 'F': 22, 'G': 14, 'H': 14
        }

        for day in range(5):
            day_name = WEEKDAYS_FULL[day]
            ws = wb.create_sheet(title=day_name)
            date = self.week_start + timedelta(days=day)
            date_str = date.strftime('%Y-%m-%d')

            ws.merge_cells('A1:H1')
            ws['A1'] = f"Дата {date_str} {day_name.upper()}"
            ws['A1'].font = Font(bold=True, size=14, name='Calibri')
            ws['A1'].alignment = center_align

            headers = ['Операционная', 'Порядок операций', 'ФИО, возраст', 'Диагноз',
                       'Наименование операции', 'Хирург', 'Наркоз / М/а', 'Примечания']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=h)
                cell.font = header_font
                cell.alignment = center_align
                cell.border = header_bottom_border

            for col_letter, width in col_widths.items():
                ws.column_dimensions[col_letter].width = width

            current_row = 4

            def add_block(room_label, patients, narcosis_type, surgeon_name=None):
                nonlocal current_row
                if not patients:
                    return
                first = True
                for idx, p in enumerate(patients, 1):
                    r = current_row
                    if first:
                        cell_room = ws.cell(row=r, column=1, value=room_label)
                        cell_room.font = room_font
                        cell_room.alignment = center_align
                        first = False
                    ws.cell(row=r, column=2, value=idx).font = data_font
                    ws.cell(row=r, column=2).alignment = center_align

                    display = patient_parser.build_display_name(p['name'], self.surname_counts)
                    age_str = f"{p['age']} {p['age_unit']}" if p['age'] is not None else ""
                    ws.cell(row=r, column=3, value=f"{display} {age_str}".strip()).font = data_font
                    ws.cell(row=r, column=3).alignment = wrap_align

                    ws.cell(row=r, column=4, value=p['diagnosis']).font = data_font
                    ws.cell(row=r, column=4).alignment = wrap_align

                    ws.cell(row=r, column=5, value=p['operation']).font = data_font
                    ws.cell(row=r, column=5).alignment = wrap_align

                    surgeon = surgeon_name if surgeon_name else p.get('surgeon', '')
                    ws.cell(row=r, column=6, value=surgeon).font = data_font
                    ws.cell(row=r, column=6).alignment = center_align

                    ws.cell(row=r, column=7, value=narcosis_type).font = data_font
                    ws.cell(row=r, column=7).alignment = center_align

                    ws.cell(row=r, column=8, value="").font = data_font
                    ws.cell(row=r, column=8).alignment = center_align

                    current_row += 1
                current_row += 1

            add_block("5", self.daily_blocks[day].get("5", []), "ЭТН", config_surgeons.SURGEON_5[day])
            add_block("7", self.daily_blocks[day].get("7", []), "М/А", config_surgeons.SURGEON_7)
            ma_patients = self.daily_blocks[day].get("MA", [])
            ma_surgeon = ma_patients[0]['surgeon'] if ma_patients else None
            add_block("М/А", ma_patients, "М/А", ma_surgeon)

        # Лист «Поступление» (перед статистикой)
        self._write_admission_sheet(wb.create_sheet(title="Поступление"))

        # Лист «Статистика»
        ws_stats = wb.create_sheet("Статистика")
        op_counter = Counter()
        surgeon_counter = Counter()
        age_groups = {"0-14": 0, "15-17": 0, "18-64": 0, "65+": 0}
        for day in range(5):
            for room in ["5", "7", "MA"]:
                for p in self.daily_blocks[day][room]:
                    ops = [o.strip().title() for o in p['operation'].split(',')]
                    for op in ops:
                        op_counter[op] += 1
                    surgeon = p.get('surgeon', 'Не указан')
                    surgeon_counter[surgeon] += 1
                    age = p.get('age')
                    if age is not None:
                        if age <= 14:
                            age_groups["0-14"] += 1
                        elif age <= 17:
                            age_groups["15-17"] += 1
                        elif age <= 64:
                            age_groups["18-64"] += 1
                        else:
                            age_groups["65+"] += 1

        ws_stats.append(["Статистика операций"])
        ws_stats.append(["Операция", "Количество"])
        for op, cnt in op_counter.most_common():
            ws_stats.append([op, cnt])

        ws_stats.append([])
        ws_stats.append(["Хирург", "Количество операций"])
        for surg, cnt in surgeon_counter.most_common():
            ws_stats.append([surg, cnt])

        ws_stats.append([])
        ws_stats.append(["Возрастная группа", "Количество пациентов"])
        for grp, cnt in age_groups.items():
            ws_stats.append([grp, cnt])

        # Автоподбор ширины столбцов
        for col_cells in ws_stats.columns:
            max_length = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 4
            ws_stats.column_dimensions[col_letter].width = adjusted_width

        wb.save(output_path)
        return True

    def _write_admission_sheet(self, ws_adm):
        """Заполняет лист «Поступление» (те же данные, что в полном плане)."""
        wrap_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        adm_header_font = Font(bold=True, size=11, name='Calibri')
        adm_data_font = Font(size=12, name='Calibri')
        small_font = Font(size=10, name='Calibri')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        adm_headers = [
            '№ п/п', 'День недели', 'Дата поступления', 'ФИО, возраст', 'Диагноз', 'Примечания'
        ]
        notes_header = adm_headers[-1]
        for col, h in enumerate(adm_headers, 1):
            cell = ws_adm.cell(row=1, column=col, value=h)
            cell.font = adm_header_font
            cell.border = thin_border
            cell.alignment = center_align

        ws_adm.column_dimensions['A'].width = 6
        ws_adm.column_dimensions['B'].width = 8
        ws_adm.column_dimensions['C'].width = 14
        ws_adm.column_dimensions['D'].width = 26
        ws_adm.column_dimensions['E'].width = 32
        ws_adm.column_dimensions['F'].width = len(notes_header) + 1

        all_patients = []
        for day in range(5):
            for room, narc_type in [("5", "ЭТН"), ("7", "М/А"), ("MA", "М/А")]:
                for p in self.daily_blocks[day][room]:
                    p['day_of_week'] = day
                    p['narc_type'] = narc_type
                    all_patients.append(p)

        prev_adm_day = None
        row_adm = 2
        for idx, p in enumerate(all_patients, 1):
            day_idx = p['day_of_week']
            operation_date = self.week_start + timedelta(days=day_idx)
            admission_date = operation_date - timedelta(days=1)
            adm_day_idx = admission_date.weekday()
            day_short = WEEKDAYS_RU[adm_day_idx]
            date_adm_str = admission_date.strftime('%d.%m.%Y')
            display = patient_parser.build_display_name(p['name'], self.surname_counts)
            age_str = f"{p['age']} {p['age_unit']}" if p['age'] is not None else ""

            if prev_adm_day is not None and adm_day_idx != prev_adm_day:
                for col in range(1, 7):
                    ws_adm.cell(row=row_adm, column=col).border = Border(top=Side(style='thin'))
            prev_adm_day = adm_day_idx

            ws_adm.cell(row=row_adm, column=1, value=idx).font = small_font
            ws_adm.cell(row=row_adm, column=1).alignment = center_align

            ws_adm.cell(row=row_adm, column=2, value=day_short).font = small_font
            ws_adm.cell(row=row_adm, column=2).alignment = center_align

            ws_adm.cell(row=row_adm, column=3, value=date_adm_str).font = small_font
            ws_adm.cell(row=row_adm, column=3).alignment = center_align

            cell_fio = ws_adm.cell(row=row_adm, column=4, value=f"{display} {age_str}".strip())
            cell_fio.font = adm_data_font
            cell_fio.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            cell_diag = ws_adm.cell(
                row=row_adm, column=5, value=f"{p['diagnosis']} {p['narc_type']}"
            )
            cell_diag.font = small_font
            cell_diag.alignment = wrap_align

            cell_notes = ws_adm.cell(row=row_adm, column=6, value=p.get("notes") or "")
            cell_notes.font = small_font
            cell_notes.alignment = wrap_align

            row_adm += 1

        ws_adm.print_gridlines = True
        ws_adm.page_setup.orientation = 'portrait'
        ws_adm.page_setup.paperSize = 9
        ws_adm.page_setup.fitToWidth = 1
        ws_adm.page_setup.fitToHeight = 1
        ws_adm.page_setup.fitToPage = True
        ws_adm.page_margins.left = 0.5
        ws_adm.page_margins.right = 0.5
        ws_adm.page_margins.top = 0.5
        ws_adm.page_margins.bottom = 0.5

    def generate_admissions_excel(self, output_path):
        """Отдельный Excel только с листом «Поступление»."""
        if self.week_start is None:
            raise ValueError("Не определена дата начала недели")
        wb = Workbook()
        ws = wb.active
        ws.title = "Поступление"
        self._write_admission_sheet(ws)
        wb.save(output_path)
        return True
