"""Служебные пометки → Примечания на листе Поступление."""

from openpyxl import load_workbook

from plan_core import OperationPlanGenerator
from room_rules import (
    classify_calendar_title,
    extract_service_notes,
    is_pure_note_event,
    is_service_event,
    reload_room_rules,
)


MONDAY = "29.06.2026"


def event(title, time, desc=""):
    return {
        "Название события": title,
        "Описание": desc,
        "Дата начала (МСК)": MONDAY,
        "Время начала (МСК)": time,
    }


def setup_function():
    reload_room_rules()


def test_extract_notes_from_patient_title():
    cleaned, notes = extract_service_notes("Иванов 30 септо перенос")
    assert "перенос" not in cleaned.lower()
    assert "Иванов" in cleaned
    assert "септо" in cleaned.lower()
    assert notes == "Перенос"


def test_extract_da_word_boundary_not_substring():
    cleaned, notes = extract_service_notes("Иванов 30 септо ДА")
    assert notes == "ДА"
    assert "септо" in cleaned.lower()
    # «да» внутри других слов не трогаем
    cleaned2, notes2 = extract_service_notes("Иванов 30 септопластика")
    assert notes2 == ""
    assert "септопластика" in cleaned2.lower()


def test_pure_note_events_classified():
    assert is_pure_note_event("перенос")
    assert is_pure_note_event("ДА")
    assert is_pure_note_event("Джабраил")
    assert classify_calendar_title("Перенос") == "service_note"
    assert classify_calendar_title("ДА") == "service_note"
    assert classify_calendar_title("Джабраил") == "service_note"
    assert is_service_event("ДА")
    assert classify_calendar_title("Иванов 30 септо перенос") is None


def test_notes_from_title_land_on_patient():
    gen = OperationPlanGenerator(
        events_data=[event("Петров 30 септо перенос", "10:00")]
    )
    gen.parse_all_events()
    gen.distribute_patients()
    patients = gen.daily_blocks[0]["5"]
    assert len(patients) == 1
    assert patients[0]["notes"] == "Перенос"
    assert "перенос" not in (patients[0].get("diagnosis_raw") or "").lower()


def test_standalone_note_attaches_to_nearest_patient():
    messages = []
    gen = OperationPlanGenerator(
        events_data=[
            event("Иванов 30 септо", "10:00"),
            event("Джабраил", "10:05"),
            event("Сидоров 25 адено", "14:00"),
        ],
        log_callback=lambda m, tag="info": messages.append(m),
    )
    gen.parse_all_events()
    gen.distribute_patients()
    ivanov = next(p for p in gen.daily_blocks[0]["5"] if "Иванов" in p["name"])
    sidorov = next(p for p in gen.daily_blocks[0]["5"] if "Сидоров" in p["name"])
    assert "Джабраил" in ivanov.get("notes", "")
    assert "Джабраил" not in sidorov.get("notes", "")


def test_admission_sheet_standalone_has_no_notes_column(tmp_path):
    gen = OperationPlanGenerator(
        events_data=[
            event("Иванов 30 септо ДА", "10:00"),
            event("перенос", "10:10"),
        ]
    )
    gen.parse_all_events()
    gen.distribute_patients()
    gen.assign_surgeons()
    gen.sort_patients_in_rooms()
    out = tmp_path / "adm.xlsx"
    gen.generate_admissions_excel(out)
    ws = load_workbook(out)["Поступление"]
    assert ws["E1"].value == "Диагноз"
    assert ws["F1"].value is None  # отдельный файл — без «Примечания»
    assert ws["E2"].font.size == 10
    patients = gen.daily_blocks[0]["5"]
    assert any("ДА" in (p.get("notes") or "") for p in patients)


def test_full_plan_admission_sheet_has_notes(tmp_path):
    gen = OperationPlanGenerator(
        events_data=[event("Иванов 30 септо Джабраил", "10:00")]
    )
    gen.parse_all_events()
    gen.distribute_patients()
    gen.assign_surgeons()
    gen.sort_patients_in_rooms()
    out = tmp_path / "plan.xlsx"
    gen.generate_excel(out)
    ws = load_workbook(out)["Понедельник"]
    assert ws["H3"].value == "Примечания"
    assert not (ws["H4"].value or "").strip()
    # в полном плане на листе «Поступление» примечания есть
    adm = load_workbook(out)["Поступление"]
    assert adm["F1"].value == "Примечания"
    assert "Джабраил" in (adm["F2"].value or "")


def test_weekday_sheet_notes_stay_empty(tmp_path):
    gen = OperationPlanGenerator(
        events_data=[event("Иванов 30 септо Джабраил", "10:00")]
    )
    gen.parse_all_events()
    gen.distribute_patients()
    gen.assign_surgeons()
    gen.sort_patients_in_rooms()
    out = tmp_path / "plan.xlsx"
    gen.generate_excel(out)
    ws = load_workbook(out)["Понедельник"]
    assert ws["H3"].value == "Примечания"
    assert not (ws["H4"].value or "").strip()


def test_bracket_notes_without_diag_key_land_on_patient(tmp_path):
    gen = OperationPlanGenerator(
        events_data=[event("Абдурахманов 4 г АТ (санация рта)", "10:00")]
    )
    gen.parse_all_events()
    gen.distribute_patients()

    notes = []
    for day in range(5):
        for room in ("5", "7", "MA"):
            for p in gen.daily_blocks[day][room]:
                if p.get("notes"):
                    notes.append(p["notes"])

    assert any("Санация рта" in n for n in notes)


def test_bracket_notes_with_diag_key_are_ignored(tmp_path):
    gen = OperationPlanGenerator(
        events_data=[event("Никитин 42 септо (адено?)", "10:00")]
    )
    gen.parse_all_events()
    gen.distribute_patients()

    patient = None
    for day in range(5):
        for room in ("5", "7", "MA"):
            for p in gen.daily_blocks[day][room]:
                if "Никитин" in p.get("name", ""):
                    patient = p
                    break
            if patient:
                break
        if patient:
            break

    assert patient is not None
    assert not (patient.get("notes") or "").strip()


def test_extract_service_notes_from_дж_bracket():
    cleaned, notes = extract_service_notes("Иванов 30 септо (дж)")
    # По умолчанию «дж» → отдельный токен (пользователь может переопределить через Review).
    assert notes.strip() == "ДЖ"
    assert "септо" in cleaned.lower()


def test_transfer_place_is_service_note_not_patient():
    gen = OperationPlanGenerator(
        events_data=[
            event("Иванов 30 септо ДА", "10:00"),
            event("Место для переноса с 04.09.2026", "10:05"),
        ]
    )
    gen.parse_all_events()
    gen.distribute_patients()

    names = []
    all_notes = []
    for day in range(5):
        for room in ("5", "7", "MA"):
            for p in gen.daily_blocks[day][room]:
                names.append(p.get("name", ""))
                if p.get("notes"):
                    all_notes.append(p["notes"])

    assert not any("Место" in n for n in names)
    assert any("Перенос" in n for n in all_notes)
