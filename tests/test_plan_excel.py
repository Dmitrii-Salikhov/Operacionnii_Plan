from openpyxl import load_workbook
from datetime import date

from constants import WEEKDAYS_FULL
from plan_core import OperationPlanGenerator, admissions_excel_filename


def test_generate_excel_creates_weekday_and_admission_sheets(tmp_path):
    generator = OperationPlanGenerator(
        events_data=[
            {
                "Название события": "Иванов Иван 30 септо",
                "Дата начала (МСК)": "29.06.2026",
                "Время начала (МСК)": "10:00",
            },
            {
                "Название события": "Петров 8 адено",
                "Дата начала (МСК)": "30.06.2026",
                "Время начала (МСК)": "11:00",
            },
        ]
    )
    generator.parse_all_events()
    generator.distribute_patients()
    generator.assign_surgeons()
    generator.sort_patients_in_rooms()
    output = tmp_path / "plan.xlsx"
    assert generator.generate_excel(output) is True
    assert output.stat().st_size > 0
    workbook = load_workbook(output)
    assert set(WEEKDAYS_FULL).issubset(workbook.sheetnames)
    assert "Поступление" in workbook.sheetnames


def test_generate_admissions_excel_only_admission_sheet(tmp_path):
    generator = OperationPlanGenerator(
        events_data=[
            {
                "Название события": "Иванов Иван 30 септо",
                "Дата начала (МСК)": "29.06.2026",
                "Время начала (МСК)": "10:00",
            },
        ]
    )
    generator.parse_all_events()
    generator.distribute_patients()
    generator.assign_surgeons()
    generator.sort_patients_in_rooms()
    output = tmp_path / admissions_excel_filename(generator.week_start)
    assert generator.generate_admissions_excel(output) is True
    workbook = load_workbook(output)
    assert workbook.sheetnames == ["Поступление"]
    assert workbook["Поступление"]["A1"].value == "№ п/п"
    assert workbook["Поступление"]["D2"].value  # ФИО


def test_admissions_excel_filename_format():
    name = admissions_excel_filename(date(2026, 6, 29))
    assert name == "Список поступлений ЛОР с 29.06.2026 по 05.07.2026.xlsx"
