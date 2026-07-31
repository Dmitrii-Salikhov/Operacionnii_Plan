"""Хирурги из UI должны попадать в Excel после surgeons.save."""

from openpyxl import load_workbook

import config_surgeons
from bridge import handlers
from plan_core import OperationPlanGenerator


def test_surgeons_save_updates_excel_column(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    monkeypatch.setenv("PLAN_BASE_DIR", str(tmp_path))
    monkeypatch.setattr(config_surgeons, "SURGEON_CONFIG_FILE", str(tmp_path / "surgeons.json"))

    # Старые значения в памяти процесса (как после старта приложения).
    old_map = {d: "Старый Х.Х." for d in range(5)}
    config_surgeons.SURGEON_5 = dict(old_map)
    config_surgeons.SURGEON_7 = "Старый7 Н.Н."
    config_surgeons.SURGEON_MA = {d: "СтарыйМА М.М." for d in range(5)}
    config_surgeons.FORBIDDEN_MA = []

    # Сохранение через bridge (как кнопка «Сохранить» в UI).
    handlers.dispatch(
        "surgeons.save",
        {
            "surgeon_5": {str(d): "Новый5 А.А." for d in range(5)},
            "surgeon_7": "Новый7 Б.Б.",
            "surgeon_ma": {str(d): "НовыйМА В.В." for d in range(5)},
            "forbidden_ma": [],
        },
    )

    gen = OperationPlanGenerator(
        events_data=[
            {
                "Название события": "Иванов Иван 30 септо",
                "Дата начала (МСК)": "29.06.2026",
                "Время начала (МСК)": "10:00",
            },
            {
                "Название события": "Петров Пётр 40 М/А септо",
                "Дата начала (МСК)": "29.06.2026",
                "Время начала (МСК)": "11:00",
            },
        ]
    )
    gen.parse_all_events()
    gen.distribute_patients()
    gen.assign_surgeons()
    gen.sort_patients_in_rooms()
    out = tmp_path / "plan.xlsx"
    assert gen.generate_excel(out) is True

    wb = load_workbook(out)
    # Понедельник: колонка F (6) — хирург
    ws = wb["Понедельник"]
    surgeons_in_sheet = []
    for row in ws.iter_rows(min_row=4, max_col=6, values_only=True):
        if row[5]:
            surgeons_in_sheet.append(row[5])

    assert "Новый5 А.А." in surgeons_in_sheet
    assert "НовыйМА В.В." in surgeons_in_sheet
    assert "Старый Х.Х." not in surgeons_in_sheet
    assert "СтарыйМА М.М." not in surgeons_in_sheet
